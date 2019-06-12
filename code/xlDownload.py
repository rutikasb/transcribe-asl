import openpyxl
import cv2
import urllib.request
import os
import re


book = openpyxl.load_workbook('../data/dai-asllvd-BU_glossing_with_variations_HS_information-extended-urls-RU.xlsx')

sheet = book.active


#EDIT THE MAX_SIGNS_TO_DOWNLOAD PARAM TO DECIDE HOW MANY ASL CATEGORIES TO DOWNLOAD
MAX_SIGNS_TO_DOWNLOAD = 50

startRow=351
glossNameCol = 2
consultantNameCol = 3
sequenceCol = 13
sceneCol = 14
frameStartCol = 15
frameEndCol = 16
baseURL = 'http://csr.bu.edu/ftp/asl/asllvd/asl-data2/quicktime/'
cameraNum = "1"


def extractFrames(filename):
	print("Extracting frames in " + filename)
	parts = os.path.split(filename)
	glossname = os.path.split(parts[0])[-1]
	name = parts[1]
	p = re.findall(r"_start\d+",name)
	startFrame = int(re.findall(r"\d+",p[0])[0])
	p = re.findall(r"_end\d+",name)
	endFrame = int(re.findall(r"\d+",p[0])[0])

	cap = cv2.VideoCapture(filename)
	total_frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
	width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
	height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
	fps = int(cap.get(cv2.CAP_PROP_FPS))

	outFolder = os.path.join("../data/videos_cropped", glossname)
	if not os.path.exists(outFolder):
		os.makedirs(outFolder)

	frameNum = startFrame
	cap.set(cv2.CAP_PROP_POS_FRAMES, startFrame)
	outFile = os.path.join(outFolder, name)
	fourcc = cv2.VideoWriter_fourcc(*"mp4v")
	# fourcc = cv2.VideoWriter_fourcc('H', '2', '6', '4')
	outcap = cv2.VideoWriter(outFile,fourcc, fps, (width,height))
	while(cap.isOpened() and frameNum < endFrame):
		ret, frame = cap.read()
		cv2.imshow('frame',frame)
		if cv2.waitKey(10) & 0xFF == ord('q'):
			break
		outcap.write(frame)
		frameNum += 1

	cap.release()
	cv2.destroyAllWindows()
	print("Done " + filename)



def downloadVideos():
	GlossDownloaded = 0
	r = startRow
	while GlossDownloaded < MAX_SIGNS_TO_DOWNLOAD:
		v = sheet.cell(row=r, column=glossNameCol)
		if(v.value is not None):
			print("\nMain New Gloss: " + v.value)
			videoFolderPath = "../data/videos/" + str(v.value).replace("/",'-')
			if not os.path.exists(videoFolderPath):
				os.makedirs(videoFolderPath)

			print("---------------------")
			i=r+1
			validRows = []
			while True:
				vv = sheet.cell(row=i, column=consultantNameCol)
				if(vv.value == "============"):
					r=i
					break
				else:
					if (vv.value is not None) and (vv.value != '------------'):
						validRows.append(i)
				i += 1
			for i in validRows:
				seq = sheet.cell(row=i, column=sequenceCol)
				scene = sheet.cell(row=i, column=sceneCol)
				frameStart = sheet.cell(row=i, column=frameStartCol)
				frameEnd = sheet.cell(row=i, column=frameEndCol)
				st = str(seq.value) + "/" + "scene" + str(scene.value) + "-" + "camera" + cameraNum + ".mov"
				url = baseURL + st
				print(url)
				filename = str(seq.value) + "-" + "scene" + str(scene.value) + "-" + "camera" + cameraNum + "_start" + str(frameStart.value) + "_" + "end" + str(frameEnd.value) + ".mov"
				filePath = os.path.join(videoFolderPath, filename)
				urllib.request.urlretrieve(url, filePath)
				extractFrames(filePath)
			GlossDownloaded += 1


if __name__ == '__main__':
	downloadVideos()

	# filename = "../data/videos/TWENTY/ASL_2008_05_12a-scene1-camera1_start2400_end2480.mov"
	# extractFrames(filename)